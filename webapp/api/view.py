# _*_ coding: utf-8 _*_
from . import _api

from werkzeug.utils import secure_filename
from flask import render_template, request, send_from_directory, abort, flash, redirect, send_file
from flask_login import login_required, current_user
import os
import re
import zipfile
import xlrd
import pymysql
from pypinyin import lazy_pinyin
from .form import UploadForm, excels, DownloadForm
from .. import conn
import pandas as pd
from openpyxl import load_workbook
from ..models import BaobiaoToSet

pardir = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
basedir = os.path.abspath(os.path.dirname(__file__))
ALLOWED_EXTENSIONS = set(['xlsx'])


# 用于判断文件后缀
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# 获取数据库中报表名
def get_baobiao_name():
    result = BaobiaoToSet.query.order_by(BaobiaoToSet.id).all()
    FILE_TO_SET = {}
    for i in range(len(result)):
        rs = str(result[i]).split(',')
        file = str(rs[0].strip('"').strip("'"))
        freq = str(rs[1].strip('"').strip("'"))
        FILE_TO_SET[str(i+1)] = file
    return FILE_TO_SET


def get_baobiao_freq():
    result = BaobiaoToSet.query.order_by(BaobiaoToSet.id).all()
    FREQ_OF_FILE = {}
    for i in range(len(result)):
        rs = str(result[i]).split(',')
        file = str(rs[0].strip('"').strip("'"))
        freq = str(rs[1].strip('"').strip("'"))
        FREQ_OF_FILE[file] = freq
    return FREQ_OF_FILE


@_api.route('/upload/')
@login_required
def upload():
    return render_template('upload.html')


@_api.route('/upload_file/', methods=['POST'])
@login_required
def upload_file():
    if request.method == 'POST':
        # get current auth
        username = current_user.username
        # check if the post request has the file part
        filedir = os.path.join(pardir, 'Files', 'upload')
        if not os.path.exists(filedir):
            os.mkdir(filedir)
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('No file selected for uploading')
            return redirect(request.url)
        files = request.files.getlist("file")
        for file in files:
            if file and allowed_file(file.filename):
                filename_chinese = re.split('[_.]', file.filename)[0]
                filename_english = ''.join(lazy_pinyin(filename_chinese))
                filename = filename_chinese + '.xlsx'
                file.save(os.path.join(filedir, filename))
                importintodb(os.path.join(filedir, filename), filename_chinese, filename_english)
            else:
                flash('仅支持xlsx文件，此模板上传失败：'+filename_chinese)
        flash('除去上面弹出报错的模板外，若有上传其他模板，则其他模板上传成功')
        return redirect('/api/upload')


def importintodb(file_to_generate, filename_chinese, filename_english):
    FILE_TO_SET = get_baobiao_name()
    FREQ_OF_FILE = get_baobiao_freq()
    conn.ping(reconnect=True)
    # 创建table
    tablename_chinese = filename_chinese
    tablename = filename_english
    try:
        freq = FREQ_OF_FILE[tablename_chinese]
    except:
        flash('未在报表名管理中维护此报表，上传此模板失败：' + filename_chinese)
    sql = 'drop table if exists ' + tablename
    cursor = conn.cursor()
    cursor.execute(sql)
    sql = """create table {} (tablename VARCHAR(100), sheetname VARCHAR(100), position VARCHAR(100), 
            content VARCHAR(500), content_list VARCHAR(500), freq VARCHAR(10), editable Boolean, 
            primary key (sheetname, position));""".format(tablename)
    cursor.execute(sql)

    wb = load_workbook(file_to_generate)
    sheet_names = wb.get_sheet_names()
    for sheet_name in sheet_names:
        sheet_ranges = wb.get_sheet_by_name(sheet_name)
        nrows = sheet_ranges.max_row
        ncols = sheet_ranges.max_column
        if nrows == 1 and ncols == 1:
            continue
        cols = [chr(i + ord('A')) for i in range(ncols)]
        rows = [str(i + 1) for i in range(nrows)]
        df = pd.DataFrame(sheet_ranges.values)
        df = df.fillna("")
        try:
            for i in range(nrows):
                for j in range(ncols):
                    position = cols[j] + rows[i]
                    content = str(df.iloc[i, j])
                    editable = False
                    if len(content) > 0 and content[0] == "|":
                        editable = True
                    sql = """insert into {tablename} (tablename, sheetname, position, content, freq, editable) values 
                          ("{tablename_chinese}", "{sheet_name}", "{position}", "{content}", "{freq}", {editable});
                          """.format(tablename=tablename, tablename_chinese=tablename_chinese, sheet_name=str(sheet_name),
                                     position=position, content=str(content), freq=str(freq), editable=str(editable))
                    # print(sql)
                    cursor.execute(sql)
            conn.commit()
        except:
            print(sql)
            print('Import Into Table Failure')
        finally:
            pass
    conn.close()


@_api.route('/download/', methods=['GET', 'POST'])
@login_required
def download():
    FILE_TO_SET = get_baobiao_name()
    form = DownloadForm()
    form.excels.choices = [(a.id, a.file) for a in BaobiaoToSet.query.all()]
    downloadlist = request.values.getlist('excels')
    if downloadlist == []:
        return render_template('download.html', form=form)
    else:
        generatedate = request.values.get('generatedate')
        generatedate = generatedate.replace('-', '_')
        filedir = os.path.join(pardir, 'Files', 'Generate')
        downloaddir = os.path.join(pardir, 'Files', 'Download')
        downloadfilename = 'Baobiao_' + current_user.username.lower() + '.zip'
        if os.path.exists(downloaddir + '/' + downloadfilename):
            os.remove(downloaddir + '/' + downloadfilename)
        zipf = zipfile.ZipFile(downloaddir + '/' + downloadfilename, 'w', zipfile.ZIP_DEFLATED)
        for filetodownload in downloadlist:
            filefolder = FILE_TO_SET[filetodownload]
            filename = filefolder + '_' + generatedate + '.xlsx'
            print(filename)
            if os.path.isfile(os.path.join(filedir, filefolder, filename)):
                zipf.write(filedir + '/' + filefolder + '/' + filename, filename)
        zipf.close()
        return send_file(downloaddir + '/' + downloadfilename, mimetype='zip',
                         attachment_filename=downloadfilename, as_attachment=True)

