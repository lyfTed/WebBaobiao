# _*_ coding: utf-8 _*_
from . import _baobiao

import codecs
import pandas as pd
from flask import render_template, request, flash, redirect, url_for, jsonify, send_file
from pymysql import err
from flask_login import login_required, current_user
import os
import re
from .form import BaobiaoForm, TianxieForm, QueryForm, PreviewForm
from pypinyin import lazy_pinyin
from openpyxl import load_workbook
from openpyxl.styles import numbers
# from win32process import SetProcessWorkingSetSize
# from win32api import GetCurrentProcessId, OpenProcess
# from win32con import PROCESS_ALL_ACCESS
import win32com.client as win32
from datetime import datetime, date, timedelta
from .form import GenerateForm, excels
from .. import conn
from ..models import BaobiaoToSet
import pythoncom
import gc
from PIL import ImageGrab

basedir = os.path.abspath(os.path.dirname(__file__))
pardir = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
ALLOWED_EXTENSIONS = set(['xlsx'])


# 用于判断文件后缀
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS



# 获取数据库中报表名
def get_baobiao_name():
    result = BaobiaoToSet.query.order_by(BaobiaoToSet.id).all()
    FILE_TO_SET = {}
    for i in range(len(result)):
        rs = str(result[i]).split(',')
        file = str(rs[0].strip('"').strip("'"))
        FILE_TO_SET[str(i+1)] = file
    return FILE_TO_SET


def get_baobiao_freq():
    result = BaobiaoToSet.query.order_by(BaobiaoToSet.id).all()
    FREQ_OF_FILE = {}
    for i in range(len(result)):
        rs = str(result[i]).split(',')
        file = str(rs[0].strip('"').strip("'"))
        freq = str(rs[1].strip('"').strip("'"))
        FREQ_OF_FILE[file] = freq
    return FREQ_OF_FILE


def get_baobiao_auditor():
    result = BaobiaoToSet.query.order_by(BaobiaoToSet.id).all()
    AUDITOR_OF_FILE = {}
    for i in range(len(result)):
        rs = str(result[i]).split(',')
        file = str(rs[0].strip('"').strip("'"))
        auditor = str(rs[2].strip('"').strip("'"))
        AUDITOR_OF_FILE[file] = auditor
    return AUDITOR_OF_FILE


@_baobiao.route('/split/', methods=['GET', 'POST'])
@login_required
def split():
    FILE_TO_SET = get_baobiao_name()
    form = BaobiaoForm()
    form.excels.choices = [(a.id, a.file) for a in BaobiaoToSet.query.all()]
    conn.ping(reconnect=True)
    cursor = conn.cursor()
    filetosetlist = request.values.getlist("excels")
    # print(filetosetlist)
    if filetosetlist == []:
        return render_template("baobiao.html", form=form)
    else:
        split_fail_list = []
        for file in filetosetlist:
            try:
                baobiao_split(cursor, file)
                conn.commit()
            except:
                split_fail_list.append(FILE_TO_SET[file])
        if len(split_fail_list) == 0:
            flash('所有所选报表拆分成功')
        else:
            flash('以下报表未上传模板，拆分失败 ' + '，'.join(split_fail_list))
            if len(filetosetlist) != len(split_fail_list):
                flash('其他所选报表拆分成功')
        conn.close()
        return render_template('baobiao.html', form=form)


def baobiao_split(cursor, file):
    FILE_TO_SET = get_baobiao_name()
    FREQ_OF_FILE = get_baobiao_freq()
    filetoset_chinese = FILE_TO_SET[file]
    filetoset = ''.join(lazy_pinyin(FILE_TO_SET[file])).lower()
    freq = FREQ_OF_FILE[filetoset_chinese]
    sql = 'select distinct position, content, sheetname from ' + filetoset + ' where editable=True;'
    cursor.execute(sql)
    conn.commit()
    sqlresult = cursor.fetchall()
    userlist = []
    userset = {}
    lastmonthend = date(date.today().year, date.today().month, 1) - timedelta(days=1)
    lastmonth = lastmonthend.strftime("_%Y_%m")
    for i in range(len(sqlresult)):
        # 获取哪个格子
        position = sqlresult[i][0]
        sheetname = sqlresult[i][2]
        # 获取计算式子
        ###################
        content_list = re.split('[-+*/()（）]', sqlresult[i][1].lstrip('|'))
        content_list = [s for s in content_list if s != '']
        # 存储分割后的数据进报表表
        content_store = ';'.join(content_list)
        sql = 'update ' + filetoset + ' set content_list="' + str(content_store) + '" where position="' + str(position) \
              + '" and sheetname="' + str(sheetname) + '";'
        cursor.execute(sql)
        conn.commit()
        # 拆分到用户表
        for content in content_list:
            userandvalue = re.split(':|：', content)
            user = ''.join(lazy_pinyin(userandvalue[0]))
            if len(userandvalue) > 1:
                value = userandvalue[1]
            else:
                value = None
            if user not in userlist:
                userlist.append(user)
                userset[user] = []
            userset[user].append((position, value, sheetname))
    for user in userlist:
        sql = 'create table if not exists ' + user + '(baobiao VARCHAR(100), sheetname VARCHAR(100), ' \
                'position VARCHAR(100), content VARCHAR(500), value_last Double, value DOUBLE, ' \
                'submit_time VARCHAR(20), freq VARCHAR(5), content_concerned VARCHAR(500));'
        cursor.execute(sql)
        sql = 'delete from ' + user + ' where baobiao="' + filetoset_chinese + '";'
        cursor.execute(sql)
        try:
            sql = 'delete from ' + user + lastmonth + ' where baobiao="' + filetoset_chinese + '";'
            cursor.execute(sql)
        except:
            pass
        for i in range(len(userset[user])):
            # print(userset[user][i])
            position = userset[user][i][0]
            value = userset[user][i][1]
            sheetname = userset[user][i][2]
            sql = 'insert into ' + user + ' (baobiao, sheetname, position, content, freq) values ("' + filetoset_chinese\
                  + '", "' + str(sheetname) + '", "' + str(position) + '", "' + str(value) + '", "' + str(freq) + '");'
            cursor.execute(sql)
            try:
                sql = 'insert into ' + user + lastmonth + ' (baobiao, sheetname, position, content, freq) values ("' + \
                      filetoset_chinese + '", "' + str(sheetname) + '", "' + str(position) + '", "' + str(value) + '", "' + str(freq) + '");'
                cursor.execute(sql)
            except:
                pass
        sql = 'select distinct content from ' + user + ';'
        cursor.execute(sql)
        distinct_content_list = [x[0] for x in cursor.fetchall()]
        for distinct_content in distinct_content_list:
            sql = 'select distinct baobiao, sheetname, position from ' + user + ' where content="' + str(distinct_content) +'";'
            cursor.execute(sql)
            sqlresult = cursor.fetchall()
            rs = '、'.join(['-'.join(x) for x in sqlresult])
            sql = 'update ' + user + ' set content_concerned="' + rs + '" where content="' + distinct_content + '";'
            print(sql)
            cursor.execute(sql)
            try:
                sql = 'update ' + user + lastmonth + ' set content_concerned="' + rs + '" where content="' + distinct_content + '";'
                cursor.execute(sql)
            except:
                pass


def exceltopng(baobiao, querydt, output, xlApp):
    if querydt is not None:
        filedir = os.path.join(pardir, 'Files', 'generate', baobiao)
    else:
        filedir = os.path.join(pardir, 'Files', 'upload')
    destdir = os.path.join(pardir, 'static', 'img', current_user.username)
    if not os.path.exists(destdir):
        os.mkdir(destdir)
    # pythoncom.CoInitialize()
    # xlApp = win32.DispatchEx("Excel.Application")
    xlApp.Visible = False
    xlBitmap = 2
    if querydt is not None:
        excelfile = baobiao + '_' + querydt.replace('/', '_') + '.xlsx'
        exceldir = os.path.join(filedir, excelfile)
        wb = xlApp.Workbooks.Open(exceldir)
        # wb = xlApp.Workbooks.Open(filedir + '/' + baobiao + '_' + querydt.replace('/', '_') + '.xlsx')
    else:
        wb = xlApp.Workbooks.Open(filedir + '/' + baobiao + '.xlsx')
    # print(wb)
    # print(exceldir)
    ws = wb.Worksheets[1]
    # print(ws)
    ws.UsedRange.CopyPicture(Format=xlBitmap)
    img = ImageGrab.grabclipboard()
    img.save(destdir + '/' + output)
    # xlApp.Quit()


def close_excel_by_force(excel):
    import win32process
    import win32gui
    import win32api
    import win32con

    # Get the window's process id's
    hwnd = excel.Hwnd
    t, p = win32process.GetWindowThreadProcessId(hwnd)
    # Ask window nicely to close
    win32gui.PostMessage(hwnd, win32con.WM_CLOSE, 0, 0)
    # Allow some time for app to close
    time.sleep(10)
    # If the application didn't close, force close
    try:
        handle = win32api.OpenProcess(win32con.PROCESS_TERMINATE, 0, p)
        if handle:
            win32api.TerminateProcess(handle, 0)
            win32api.CloseHandle(handle)
    except:
        pass


@_baobiao.route('/fill/', methods=['GET', 'POST'])
@login_required
def fill():
    form = TianxieForm()
    # 预览填写报表
    previewform = PreviewForm()
    previewform.excel.choices = [(a.id, a.file) for a in BaobiaoToSet.query.all()]
    username = current_user.username.lower()
    conn.ping(reconnect=True)
    cursor = conn.cursor()
    lastmonthend = date(date.today().year, date.today().month, 1) - timedelta(days=1)
    lastm = lastmonthend.strftime(("%m"))
    lastmonth = lastmonthend.strftime("_%Y_%m")
    try:
        if lastm in ["01", "02", "04", "05", "07", "08", "10", "11"]:
            sql = 'create table if not exists ' + username + lastmonth + \
                  ' select distinct * from ' + username + ' where freq="M";'
            cursor.execute(sql)
            try:
                # 上一期， Monthly
                lastterm = (date(lastmonthend.year, lastmonthend.month, 1) - timedelta(days=1)).strftime("_%Y_%m")
                sql = 'update ' + username + lastmonth + ' s set s.value_last = (select value from ' + username + lastterm \
                      + ' where baobiao=s.baobiao and sheetname=s.sheetname and position=s.position and ' \
                        'content=s.content and freq="M");'
                cursor.execute(sql)
                conn.commit()
            except:
                pass
            # 不考虑不同报表和格子位置，只要填写内容即只展现一次
            sql = 'select *, count(distinct content) from ' + username + lastmonth + ' where freq="M" group by content' \
                  ' order by baobiao asc, sheetname asc, position asc;'
            cursor.execute(sql)
            sqlresult = cursor.fetchall()
        elif lastm in ["03", "09"]:
            sql = 'create table if not exists ' + username + lastmonth + \
                  ' select distinct * from ' + username + ' where freq in ("M", "Q");'
            cursor.execute(sql)
            try:
                # 上一期, Monthly
                lastterm = (date(lastmonthend.year, lastmonthend.month, 1) - timedelta(days=1)).strftime("_%Y_%m")
                sql = 'update ' + username + lastmonth + ' s set s.value_last = (select value from ' + username + lastterm \
                      + ' where baobiao=s.baobiao and sheetname=s.sheetname and position=s.position and content=s.content and freq="M");'
                cursor.execute(sql)
                conn.commit()
            except:
                pass
            try:
                # 上一期, Quarterly
                lastterm = (date(lastmonthend.year, lastmonthend.month, 1) - timedelta(days=63)).strftime("_%Y_%m")
                sql = 'update ' + username + lastmonth + ' s set s.value_last = (select value from ' + username + lastterm \
                      + ' where baobiao=s.baobiao and sheetname=s.sheetname and position=s.position and content=s.content and freq="Q");'
                cursor.execute(sql)
                conn.commit()
            except:
                pass
            sql = 'select *, count(distinct content) from ' + username + lastmonth + \
                  ' where freq in ("M", "Q") group by content order by baobiao asc, sheetname asc, position asc;'
            cursor.execute(sql)
            sqlresult = cursor.fetchall()
        elif lastm in ["06", "12"]:
            sql = 'create table if not exists ' + username + lastmonth + \
                  ' select distinct * from ' + username + ';'
            cursor.execute(sql)
            try:
                # 上一期, Monthly
                lastterm = (date(lastmonthend.year, lastmonthend.month, 1) - timedelta(days=1)).strftime("_%Y_%m")
                sql = 'update ' + username + lastmonth + ' s set s.value_last = (select value from ' + username + lastterm \
                      + ' where baobiao=s.baobiao and sheetname=s.sheetname and position=s.position and content=s.content and freq="M");'
                cursor.execute(sql)
                conn.commit()
            except:
                pass
            try:
                # 上一期, Quarterly
                lastterm = (date(lastmonthend.year, lastmonthend.month, 1) - timedelta(days=63)).strftime("_%Y_%m")
                sql = 'update ' + username + lastmonth + ' s set s.value_last = (select value from ' + username + lastterm \
                      + ' where baobiao=s.baobiao and sheetname=s.sheetname and position=s.position and content=s.content and freq="Q");'
                cursor.execute(sql)
                conn.commit()
            except:
                pass
            try:
                # 上一期, Half year
                lastterm = (date(lastmonthend.year, lastmonthend.month, 1) - timedelta(days=153)).strftime("_%Y_%m")
                sql = 'update ' + username + lastmonth + ' s set s.value_last = (select value from ' + username + lastterm \
                      + ' where baobiao=s.baobiao and sheetname=s.sheetname and position=s.position and content=s.content and freq="H");'
                cursor.execute(sql)
                conn.commit()
            except:
                pass
            sql = 'select *, count(distinct content) from ' + username + lastmonth + ' group by content order by ' \
                  'baobiao asc, sheetname asc, position asc;'
            cursor.execute(sql)
            sqlresult = cursor.fetchall()
    except:
        sqlresult = None
    if request.method == 'GET' and sqlresult is not None:
        conn.close()
        return render_template("baobiao_tianxie.html", form=form, previewform=previewform, sqlresult=sqlresult)
    elif request.method == 'GET' and sqlresult is None:
        conn.close()
        return render_template("baobiao_tianxie.html", form=form, previewform=previewform)
    elif request.method == 'POST':
        # 预览填写报表
        if previewform.preview.data:
            try:
                FILE_TO_SET = get_baobiao_name()
                baobiao = FILE_TO_SET[str(previewform.excel.data)]
                pythoncom.CoInitialize()
                xlApp = win32.DispatchEx("Excel.Application")
                exceltopng(baobiao, None, 'fill_img.png', xlApp)
                previewimg = url_for('static', filename='img/' + current_user.username + '/fill_img.png')
                # quit excel and release com object
                xlApp.Quit()
                xlApp = None
                pythoncom.CoUninitialize()
                # pid = GetCurrentProcessId()
                # handle = OpenProcess(PROCESS_ALL_ACCESS, True, pid)
                # SetProcessWorkingSetSize(handle,-1,-1)
                gc.collect()
                return render_template("preview.html", baobiao=baobiao, previewimg=previewimg)
            except:
                gc.collect()
                xlApp.Quit()
                xlApp = None
                pythoncom.CoUninitialize()
                flash('该报表模板尚未上传，请联系管理员上传')
                return redirect('/baobiao/fill')
        elif previewform.download.data:
            try:
                FILE_TO_SET = get_baobiao_name()
                baobiao = FILE_TO_SET[str(previewform.excel.data)]
                filedir = os.path.join(pardir, 'Files', 'upload')
                filepath = filedir + '/' + baobiao + '.xlsx'
                return send_file(filepath, mimetype='xlsx', as_attachment=True)
            except:
                flash('该报表模板尚未上传，请联系管理员上传')
                return redirect('/baobiao/fill')
        # 填写内容
        tianxie = request.form.getlist("values")
        try:
            for i in range(len(tianxie)):
                content = sqlresult[i][3]
                value = str(tianxie[i])
                submit_time = datetime.today().strftime('%y/%m/%d %H:%M')
                sql = 'update ' + username + lastmonth + ' set value=' + value + ' where content="' + content + '";'
                sql2 = 'update ' + username + lastmonth + ' set submit_time="' + submit_time + '" where content="' + content + '";'
                if value != '':
                    cursor.execute(sql)
                    cursor.execute(sql2)
            conn.commit()
            conn.close()
            flash("数据提交成功")
            return redirect('/baobiao/fill/')
        except:
            flash("有个格子不是数字格式，只能填写数字格式")
            return redirect('/baobiao/fill/')
        finally:
            return redirect('/baobiao/fill/')


@_baobiao.route('/upload_data/', methods=['GET', 'POST'])
@login_required
def upload_data():
    if request.method == 'GET':
        return render_template("baobiao_uploaddata.html")
    elif request.method == 'POST':
        # get current auth
        username = current_user.username
        # check if the post request has the file part
        filedir = os.path.join(pardir, 'Files', 'data', username)
        if not os.path.exists(filedir):
            os.mkdir(filedir)
        if 'data' not in request.files:
            flash('No data part')
            return redirect(request.url)
        file = request.files['data']
        if file.filename == '':
            flash('No file selected for uploading')
            return redirect(request.url)
        files = request.files.getlist("data")
        for file in files:
            if file and allowed_file(file.filename):
                filename_chinese = re.split('[_.]', file.filename)[0]
                filename_english = ''.join(lazy_pinyin(filename_chinese))
                filename = filename_chinese + '.xlsx'
                file.save(os.path.join(filedir, filename))
                try:
                    update_db_data(os.path.join(filedir, filename), filename_chinese, filename_english)
                    os.remove(os.path.join(filedir, filename))
                except:
                    flash('上传数据表中有错，上传失败')
                    pass
        flash('数据上传成功,请回到填写报表数据页面检查上传结果，若有问题，请检查上传的数据文件')
        return redirect('/baobiao/upload_data')


def update_db_data(file_to_generate, filename_chinese, filename_english):
    FILE_TO_SET = get_baobiao_name()
    FREQ_OF_FILE = get_baobiao_freq()
    conn.ping(reconnect=True)
    # 查询table
    # 拿到数据excel表名
    tablename_chinese = filename_chinese
    tablename = filename_english
    freq = FREQ_OF_FILE[tablename_chinese]
    # 获取当前用户，及填写月份
    username = current_user.username.lower()
    lastmonthend = date(date.today().year, date.today().month, 1) - timedelta(days=1)
    # lastm = lastmonthend.strftime(("%m"))
    lastmonth = lastmonthend.strftime("_%Y_%m")
    # 连上数据库
    cursor = conn.cursor()
    sql = 'select sheetname, position from ' + username + lastmonth + ' where baobiao="' + tablename_chinese + '" group by ' \
           'sheetname, position having count(*)=1;'
    cursor.execute(sql)
    conn.commit()
    sqlresult = cursor.fetchall()
    # 先查出有几个格子里的填写内容是只需要一个数据的，再去excel里找对应格子拿数
    wb = load_workbook(file_to_generate, data_only=True)
    sheet_names = wb.get_sheet_names()
    for sheetposition in sqlresult:
        sheetname = sheetposition[0]
        gezi = sheetposition[1]
        if sheetname in sheet_names:
            sheet_ranges = wb.get_sheet_by_name(sheetname)
            if pd.isnull(sheet_ranges[gezi].value):
                position_value = 0
            else:
                position_value = round(sheet_ranges[gezi].value, 2)
            submit_time = datetime.today().strftime('%y/%m/%d %H:%M')
            sql = 'update ' + username + lastmonth + ' set value=' + str(position_value) + ', submit_time="' + \
                  submit_time + '" where baobiao="' + tablename_chinese + '" and sheetname="' + sheetname + \
                  '" and position="' + gezi + '";'
            # print(sql)
            cursor.execute(sql)
            conn.commit()
    conn.close()


@_baobiao.route('/generate/', methods=['GET', 'POST'])
@login_required
def generate():
    form = GenerateForm()
    lastmonthend = date(date.today().year, date.today().month, 1) - timedelta(days=1)
    lastm = lastmonthend.strftime(("%m"))
    generate_choice_list = []
    for a in BaobiaoToSet.query.all():
        if a.auditor == current_user.username or current_user.username.lower() == 'admin':
            if lastm in ["01", "02", "04", "05", "07", "08", "10", "11"] and a.freq == 'M':
                generate_choice_list.append((a.id, a.file))
            elif lastm in ["03", "09"] and (a.freq == 'M' or a.freq == 'Q'):
                generate_choice_list.append((a.id, a.file))
            elif lastm in ["06", "12"]:
                generate_choice_list.append((a.id, a.file))
    form.excels.choices = generate_choice_list
    FILE_TO_SET = get_baobiao_name()

    generatelist = request.values.getlist('excels')
    filedir = os.path.join(pardir, 'Files', 'generate')
    if not os.path.exists(filedir):
        os.mkdir(filedir)
    if generatelist == []:
        return render_template('generate.html', form=form)
    else:
        lastmonthend = date(date.today().year, date.today().month, 1) - timedelta(days=1)
        lastmonth = lastmonthend.strftime("%Y_%m_%d")
        print(lastmonth)
        generatedate = lastmonth
        allcomplete = True
        generate_fail_list = []
        pythoncom.CoInitialize()
        xlApp = win32.DispatchEx("Excel.Application")
        xlApp.Visible = False
        xlApp.DisplayAlerts = False
        conn.ping(reconnect=True)
        for generatefile in generatelist:
            filetogenerate_chinese = FILE_TO_SET[generatefile]
            try:
                alert = generateFile(filetogenerate_chinese, generatedate, xlApp)
                if len(alert) != 0:
                    allcomplete = False
                    alertmsg = filetogenerate_chinese + ': ' + ','.join(alert)
                    flash('以下用户还未完成对应报表：' + alertmsg)
            except:
                generate_fail_list.append(filetogenerate_chinese)
        if len(generate_fail_list) != 0:
            flash('以下报表生成失败，可能是管理员尚未上传模板：' + ','.join(generate_fail_list))
        if allcomplete and len(generate_fail_list) == 0:
            flash('所选报表均生成成功')
        elif allcomplete and len(generate_fail_list) != len(generatelist):
            flash('其他所选报表生成成功')
        else:
            flash('除生成失败的报表外，其他报表生成成功，但有用户尚未完成填写，生成的报表数据还未完整，已删除')
        conn.close()
        xlApp.Quit()
        xlApp = None
        pythoncom.CoUninitialize()    
        return render_template('generate.html', form=form)


def generateFile(filetogenerate_chinese, generatedate, xlApp):
    # pythoncom.CoInitialize()
    # conn.ping(reconnect=True)
    cursor = conn.cursor()
    filetogenerate = ''.join(lazy_pinyin(filetogenerate_chinese))
    tablenamenew = filetogenerate + '_' + generatedate
    print(filetogenerate)
    print(tablenamenew)
    userlist = []
    # 获取上个月日期
    lastmonthend = date(date.today().year, date.today().month, 1) - timedelta(days=1)
    lastmonth = lastmonthend.strftime("_%Y_%m")
    # 创建新表
    sql = 'drop table if exists ' + tablenamenew
    cursor.execute(sql)
    sql = 'create table ' + tablenamenew + '(tablename VARCHAR(100), sheetname VARCHAR(100), position VARCHAR(100), ' \
        'content VARCHAR(500), editable Boolean, content_formula VARCHAR(500), content_list VARCHAR(500),' \
        ' primary key (sheetname, position));'
    print(sql)
    cursor.execute(sql)
    conn.commit()
    try:
        sql = 'insert into ' + tablenamenew + ' (tablename, sheetname, position, content, editable, content_formula, content_list) ' \
              'select tablename, sheetname, position, content, editable, content, content_list from ' + filetogenerate + ';'
        print(sql)
        cursor.execute(sql)
        conn.commit()
    except:
        print('Update Table Failure')
    finally:
        pass

    # 从模板拿需要填写的格子
    sql = 'select distinct sheetname, position, content, content_list from ' + filetogenerate + ' where editable=True;'
    cursor.execute(sql)
    # conn.commit()
    sqlresult = cursor.fetchall()
    # 用来提示哪些用户还未填写此张报表
    alertset = set()
    # 按每一个格子循环去用户表中取数
    for i in range(len(sqlresult)):
        # 获取哪个格子
        sheetname = sqlresult[i][0]
        position = sqlresult[i][1]
        formula = sqlresult[i][2]
        # 获取用户和内容
        content_list = sqlresult[i][3].split(';')
        uservalue_list = []
        for content in content_list:
            userandvalue = re.split(':|：', content)
            user = ''.join(lazy_pinyin(userandvalue[0]))
            if user not in userlist:
                userlist.append(user)
            if len(userandvalue) > 1:
                valuecontent = userandvalue[1]
            else:
                valuecontent = None
            try:
                sql = 'select value from ' + user + lastmonth + ' where baobiao="' + filetogenerate_chinese + \
                      '" and sheetname="' + str(sheetname) + '" and position="' + str(position) + '" and content="' +\
                      valuecontent + '";'
                cursor.execute(sql)
                value = cursor.fetchone()[0]
                uservalue_list.append(value)
            except err.DatabaseError:
                value = None
                uservalue_list.append(value)
            else:
                pass
            # 代入运算式
            # 需要按式子中的顺序查找对应的值，用uservalue_list代替content_list中的值并代入formula
            # print(userandvalue)
            # print(value)
            if value is None:
                alertset.add(user)
                formula = formula.replace(content, str(0))
            else:
                formula = formula.replace(content, str(value))
        formula = formula.replace("（", "(")
        formula = formula.replace("）", ")")
        # print(formula)
        try:
            positionresult = round(eval(formula.lstrip("|")), 2)
        except:
            positionresult = "=NA()" ##计算报错就在excel中填NA
        sql = 'update ' + filetogenerate + '_' + generatedate + ' set content="' + str(positionresult) + \
              '" where sheetname="' + str(sheetname) + '" and position="' + str(position) + '";'
        cursor.execute(sql)
        conn.commit()
    ######################
    # 生成excel
    # 计算行数列数
    wb = load_workbook(pardir + '/Files/upload/' + filetogenerate_chinese + '.xlsx')
    sheet_names = wb.get_sheet_names()
    for sheet_name in sheet_names:
        sh = wb.get_sheet_by_name(sheet_name)
        nrows = sh.max_row
        ncols = sh.max_column
        if nrows == 1 and ncols == 1:
            continue
        sql = 'select distinct position, content from ' + filetogenerate + '_' + generatedate + \
              ' where sheetname="' + str(sheet_name) + '" and editable=TRUE;'
        cursor.execute(sql)
        conn.commit()
        sqlresult = cursor.fetchall()
        for x in sqlresult:
            try:
                sh[x[0]] = float(x[1])
                sh[x[0]].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            except:
                sh[x[0]] = x[1]
        # 把带公式计算的格子填入公式，自动计算
        sql = 'select distinct position, content from ' + filetogenerate + ' where content like "=%";'
        cursor.execute(sql)
        conn.commit()
        sqlresult = cursor.fetchall()
        for x in sqlresult:
            sh[x[0]] = str(x[1])
            sh[x[0]].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        filedir = os.path.join(pardir, 'Files', 'Generate', filetogenerate_chinese)
        if not os.path.exists(filedir):
            os.mkdir(filedir)
        # 保存带公式的xlsx
        wb.save(filedir + '/' + filetogenerate_chinese + '_' + generatedate + '.xlsx')

    # 去除公式只保存数值,需要先excel程序打开再保存一下，然后用openpyxl只保留数值
    #### https://www.cnblogs.com/vhills/p/8327918.html
    print('Begin Open')
    xlBook = xlApp.Workbooks.Open(filedir + '/' + filetogenerate_chinese + '_' + generatedate + '.xlsx')
    print('End Open,Begin Save')
    xlBook.Save()
    print('End Save, Begin Close')
    xlBook.Close(True)
    xlBook = None
    gc.collect()
    print('End Close')
    wb = load_workbook(filedir + '/' + filetogenerate_chinese + '_' + generatedate + '.xlsx', data_only=True)
    wb.save(filedir + '/' + filetogenerate_chinese + '_' + generatedate + '.xlsx')
    print(alertset)
    if len(alertset) != 0 and os.path.exists(filedir + '/' + filetogenerate_chinese + '_' + generatedate + '.xlsx'):
        print('有用户尚未填写，删除生成的excel')
        os.remove(filedir + '/' + filetogenerate_chinese + '_' + generatedate + '.xlsx')
    return alertset


def exceltohtml(baobiao, querydt, output):
    filedir = os.path.join(pardir, 'Files', 'generate', baobiao)
    destdir = os.path.join(pardir, 'templates')
    xd = pd.ExcelFile(filedir + '/' + baobiao + '_' + querydt.replace('/', '_') + '.xlsx')
    pd.set_option('display.float_format', lambda x: format(x, ',.2f'), 'display.max_colwidth', 1000)
    df = xd.parse()
    with codecs.open(destdir + '/' + output + '.html', 'w', encoding='utf-8') as html_file:
        html_file.write(df.to_html(header=True, index=True, na_rep=''))


@_baobiao.route('/query/', methods=['GET', 'POST'])
@login_required
def query():
    form = QueryForm(form_name='QueryForm')
    form.excel.choices = [(a.id, a.file) for a in BaobiaoToSet.query.all()]
    form.querydate.choices = [(a.id, a.freq) for a in BaobiaoToSet.query.all()]
    FILE_TO_SET = get_baobiao_name()
    if request.method == 'GET':
        return render_template("baobiao_query.html", form=form)
    if request.form['form_name'] == 'QueryForm':
        baobiao = FILE_TO_SET[str(form.excel.data)]
        if form.customizeddate.data is not None:
            querydt = request.values.get('customizeddate')
            querydt = querydt.replace('-', '/')
        else:
            querydt = form.querydate.data
        try:
            pythoncom.CoInitialize()
            xlApp = win32.DispatchEx("Excel.Application")
            exceltopng(baobiao, querydt, 'query_img.png', xlApp)
            queryimg = url_for('static', filename='img/' + current_user.username + '/query_img.png')
            xlApp.Quit()
            xlApp = None
            pythoncom.CoUninitialize()
            return render_template("query.html", baobiao=baobiao, queryimg=queryimg)
        except:
            xlApp.Quit()
            flash('所选报表尚未生成')
    return redirect(url_for('baobiao.query'))


def cal_former_month(dt, N, freq):
    dtlist = []
    year = dt.year
    month = dt.month
    for i in range(N):
        if freq == 'M':
            dtlist.append((date(year, month, 1) - timedelta(days=1)).strftime("%Y/%m/%d"))
        if freq == 'Q' and i % 3 == 0:
            dtlist.append((date(year, month, 1) - timedelta(days=1)).strftime("%Y/%m/%d"))
        if freq == 'H' and i % 6 == 0:
            dtlist.append((date(year, month, 1) - timedelta(days=1)).strftime("%Y/%m/%d"))
        if month == 1:
            year -= 1
            month = 12
        else:
            month -= 1
    return(dtlist)


@_baobiao.route('/_get_freq/', methods=['GET', 'POST'])
def _get_freq():
    FILE_TO_SET = get_baobiao_name()
    FREQ_OF_FILE = get_baobiao_freq()
    excel = request.args.get('excel', default=None, type=int)
    freq = FREQ_OF_FILE[FILE_TO_SET[str(excel)]]
    query_dt = cal_former_month(date.today(), 30, freq)[0:5]
    query_dt = [(query_dt[i], query_dt[i]) for i in range(5)]
    return jsonify(query_dt)

